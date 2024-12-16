from openpyxl import load_workbook
import fitz
from utils import find_tables_pdf, find_tables_excel, clean_and_compare
import pandas as pd
from llm import extract_llm



def extract_metrics(file_url):
    """
    Non LLM method for extraction of metrics from PDF/Excel file
    :param file_url: PDF/Excel file
    :return: dataframe containing metric name with values
    """
    summary = pd.DataFrame()

    if file_url.split('.')[-1] == 'xlsx':
        # Load the workbook and access the sheet
        wb = load_workbook(file_url)
        sheet = wb['Summary']  # Replace with your sheet name

        # Extracting all rows as a table, with values only
        table_data = []
        for row in sheet.iter_rows(values_only=True):
            table_data.append(row)


        metrics = find_tables_excel(table_data)

        summary['metric_name'] = metrics.keys()
        summary['excel_extraction_Q2_24'] = [vals["2Q'24"] for key, vals in metrics.items()]
        summary['excel_extraction_Q3_24'] = [vals["3Q'24"] for key, vals in metrics.items()]


    elif file_url.split('.')[-1] == 'pdf':
        doc = fitz.open(file_url)
        page = doc.load_page(1)
        metrics = find_tables_pdf(page)

        summary['metric_name'] = metrics.keys()
        summary['pdf_extraction_Q2_24'] = [vals["2Q'24"] for key, vals in metrics.items()]
        summary['pdf_extraction_Q3_24'] = [vals["3Q'24"] for key, vals in metrics.items()]

    return summary


def extract_metrics_llm(file_url):
    """
    LLM based extraction method
    :param file_url: PDF/Excel file
    :return: dataframe containing metric name with values.
    """
    summary = pd.DataFrame()

    if file_url.split('.')[-1] == 'xlsx':

        def filter_plain_text(rows):
            plain_text = ""
            for row in rows:
                for value in row:
                    if value:
                        plain_text += str(value) + " "  # Add space between values

            return plain_text.strip()

        # Load the workbook and access the sheet
        wb = load_workbook(file_url)
        sheet = wb['Summary']  # Replace with your sheet name

        # Extracting all rows as a table, with values only
        table_data = []
        for row in sheet.iter_rows(values_only=True):
            table_data.append(row)


        excel_text = filter_plain_text(table_data)

        metrics = extract_llm(excel_text)

        summary['metric_name'] = metrics["Q2'24"].keys()
        summary['excel_extraction_Q2_24'] = metrics["Q2'24"].values()
        summary['excel_extraction_Q3_24'] = metrics["Q3'24"].values()

    elif file_url.split('.')[-1] == 'pdf':
        doc = fitz.open(file_url)
        page = doc.load_page(1)
        metrics = extract_llm(page.get_text())

        summary['metric_name'] = metrics["Q2'24"].keys()
        summary['pdf_extraction_Q2_24'] = metrics["Q2'24"].values()
        summary['pdf_extraction_Q3_24'] = metrics["Q3'24"].values()

    return summary

if __name__ == "__main__":
    pdf_url = '2024pr-qtr3rslt.pdf'
    excel_url = '3Q24-SUPP-ForWeb.xlsx'

    ### No LLM Flow
    pdf_metrics = extract_metrics(pdf_url)
    excel_metrics = extract_metrics(excel_url)

    result_csv = pd.merge(pdf_metrics,excel_metrics, on='metric_name')

    result_csv['Q2_Match'] = [clean_and_compare(row['pdf_extraction_Q2_24'], row['excel_extraction_Q2_24']) for index,row in result_csv.iterrows()]
    result_csv['Q3_Match'] = [clean_and_compare(row['pdf_extraction_Q3_24'], row['excel_extraction_Q3_24']) for
                              index, row in result_csv.iterrows()]



    ### LLM Flow
    llm_pdf_metrics = extract_metrics_llm(pdf_url)
    llm_excel_metrics = extract_metrics_llm(excel_url)

    llm_result_csv = pd.merge(llm_pdf_metrics, llm_excel_metrics, on='metric_name')

    llm_result_csv['Q2_Match'] = [clean_and_compare(row['pdf_extraction_Q2_24'], row['excel_extraction_Q2_24']) for
                              index, row in llm_result_csv.iterrows()]
    llm_result_csv['Q3_Match'] = [clean_and_compare(row['pdf_extraction_Q3_24'], row['excel_extraction_Q3_24']) for
                              index, row in llm_result_csv.iterrows()]

    result_csv.to_csv('non_llm_results.csv',index=False)
    llm_result_csv.to_csv('llm_results.csv', index=False)






